import io
import uuid
import httpx
from typing import Any, Dict, List, Optional

from app.services.media_extractor import extract_media

from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.models import (
    Distance,
    VectorParams,
    PointStruct,
    Filter,
    FieldCondition,
    MatchValue,
)

from app.config import get_settings
from app.services.embedder import embed, vector_dim, EmbedProvider

# One Qdrant collection per provider (they have different vector dimensions)
def _collection_name(provider: EmbedProvider) -> str:
    return f"staffbot_{provider}"   # staffbot_openai | staffbot_local


class DocumentIndexer:
    def __init__(self):
        settings = get_settings()
        self.qdrant = QdrantClient(url=settings.QDRANT_URL)

    # ─── Collection bootstrap ──────────────────────────────────────────────────

    def _ensure_collection(self, provider: EmbedProvider):
        name = _collection_name(provider)
        existing = {c.name for c in self.qdrant.get_collections().collections}
        if name not in existing:
            self.qdrant.create_collection(
                collection_name=name,
                vectors_config=VectorParams(
                    size=vector_dim(provider),
                    distance=Distance.COSINE,
                ),
            )
            print(f"[indexer] created collection '{name}'", flush=True)

    # ─── Public API ────────────────────────────────────────────────────────────

    async def process_document(
        self,
        document_id: str,
        tenant_id: str,
        profile_id: str,
        file_url: str,
        file_type: str,
        embed_provider: EmbedProvider = "openai",
    ) -> int:
        """
        Download, extract, chunk, embed, and upsert a document into Qdrant.
        Returns the number of chunks indexed.
        """
        self._ensure_collection(embed_provider)
        col = _collection_name(embed_provider)

        # 1. Download file
        async with httpx.AsyncClient(timeout=60) as client:
            response = await client.get(file_url)
            response.raise_for_status()
            content = response.content

        # 2. Extract text + media (images, video URLs)
        media      = extract_media(content, file_type, tenant_id, document_id)
        text       = media["text"]
        doc_images: List[Dict[str, Any]]  = media["images"]
        doc_videos: List[str]             = media["video_urls"]

        if not text.strip():
            raise ValueError("No text could be extracted from the document")

        if doc_images or doc_videos:
            print(
                f"[indexer] {document_id}: {len(doc_images)} images, {len(doc_videos)} video URLs",
                flush=True,
            )

        # 3. Split into chunks
        splitter = RecursiveCharacterTextSplitter(chunk_size=512, chunk_overlap=50)
        chunks = splitter.split_text(text)
        if not chunks:
            raise ValueError("Document produced no chunks after splitting")

        # 4. Remove old vectors for this document (re-index support)
        self.qdrant.delete(
            collection_name=col,
            points_selector=Filter(must=[
                FieldCondition(key="document_id", match=MatchValue(value=document_id)),
            ]),
        )

        # 5. Embed and upsert in batches of 64 (keeps each Qdrant payload well under 32 MB)
        total = 0
        for i in range(0, len(chunks), 64):
            batch      = chunks[i : i + 64]
            embeddings = embed(batch, embed_provider)
            points = [
                PointStruct(
                    id=str(uuid.uuid4()),
                    vector=vec,
                    payload={
                        "tenant_id":      tenant_id,
                        "profile_id":     profile_id,
                        "document_id":    document_id,
                        "embed_provider": embed_provider,
                        "chunk_index":    i + j,
                        "text":           chunk,
                        "text_preview":   chunk[:200],
                        "images":         doc_images,
                        "video_urls":     doc_videos,
                    },
                )
                for j, (chunk, vec) in enumerate(zip(batch, embeddings))
            ]
            self.qdrant.upsert(collection_name=col, points=points)
            total += len(points)

        print(f"[indexer] {document_id} ({embed_provider}): {total} chunks → {col}", flush=True)
        return total

    async def delete_document_vectors(
        self,
        document_id: str,
        tenant_id: str,
        embed_provider: EmbedProvider = "openai",
    ):
        """Remove all Qdrant points belonging to a document."""
        col = _collection_name(embed_provider)
        self.qdrant.delete(
            collection_name=col,
            points_selector=Filter(
                must=[
                    FieldCondition(key="document_id", match=MatchValue(value=document_id)),
                    FieldCondition(key="tenant_id",   match=MatchValue(value=tenant_id)),
                ]
            ),
        )

    # ─── Text extraction ───────────────────────────────────────────────────────

    def _extract_text(self, content: bytes, file_type: str) -> str:
        if file_type == "pdf":
            return self._extract_pdf(content)
        elif file_type == "docx":
            return self._extract_docx(content)
        elif file_type == "xlsx":
            return self._extract_xlsx(content)
        else:  # txt / fallback
            return content.decode("utf-8", errors="ignore")

    def _extract_pdf(self, content: bytes) -> str:
        import fitz  # pymupdf

        doc = fitz.open(stream=content, filetype="pdf")
        pages = [page.get_text() for page in doc]
        doc.close()
        return "\n\n".join(p for p in pages if p.strip())

    def _extract_docx(self, content: bytes) -> str:
        from docx import Document

        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    def _extract_xlsx(self, content: bytes) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows.append(row_text)
        wb.close()
        return "\n".join(rows)
